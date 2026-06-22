import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_oac_excel():
    # Define file paths
    output_dir = r"c:\antigravity prjs\MAS-Lean-1\managed_workspaces\ws-default-career-twin\ws-default-career-twin\outputs\proj-skillflo-pitch\03_output"
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, "OAC_Fundraising_Database_Analysis.xlsx")
    
    # Initialize workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets
    ws1 = wb.create_sheet(title="OAC_Raw_Data_Summary")
    ws2 = wb.create_sheet(title="Pitch_Deck_Metrics_Ratio")
    
    # Style definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    font_section = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_header_warn = Font(name="Calibri", size=11, bold=True, color="5B2C00")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    font_italic = Font(name="Calibri", size=9, italic=True, color="595959")
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_header_warn = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_stripe = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_accent = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_bottom_double = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=Side(style='double', color='000000'))
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Enable grid lines
    ws1.views.sheetView[0].showGridLines = True
    ws2.views.sheetView[0].showGridLines = True

    # ----------------------------------------------------
    # SHEET 1: OAC_Raw_Data_Summary
    # ----------------------------------------------------
    ws1.row_dimensions[1].height = 25
    ws1.row_dimensions[2].height = 18
    ws1.row_dimensions[3].height = 10
    
    # Title Block
    ws1["B1"] = "OAC Airtable Database Raw Summary"
    ws1["B1"].font = font_title
    ws1["B2"] = "Source: Live OAC Airtable Database Query | Data Date: 2026-06-05"
    ws1["B2"].font = font_italic
    
    # Write Table 1.1: Database Tables & Records Count (B5:C18)
    ws1["B5"] = "Table 1.1: Database Table Records Count"
    ws1["B5"].font = font_section
    
    headers_t1 = ["Table Name", "Record Count"]
    for col_idx, h in enumerate(headers_t1, start=2): # Col B and C
        cell = ws1.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
    
    t1_data = [
        ("Zoho Applications", 3823),
        ("Daily Performance Tracker", 1482),
        ("Screening", 1351),
        ("Candidates", 810),
        ("Jobs", 410),
        ("Weekly Performance Tracker", 422),
        ("Applications", 464),
        ("Clients", 72),
        ("Activity Log", 2317),
        ("Placement", 38),
        ("Users", 22),
        ("KPI and Actual", 12)
    ]
    
    for row_offset, (tbl, count) in enumerate(t1_data):
        r = 7 + row_offset
        ws1.row_dimensions[r].height = 18
        c1 = ws1.cell(row=r, column=2, value=tbl)
        c2 = ws1.cell(row=r, column=3, value=count)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        c2.font = font_regular
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '#,##0'
        
        if row_offset % 2 == 1:
            c1.fill = fill_stripe
            c2.fill = fill_stripe
            
    # Write Table 1.2: Placement & Financial Metrics (B21:E30)
    ws1["B21"] = "Table 1.2: Raw Placement & Financial Metrics"
    ws1["B21"].font = font_section
    
    headers_t2 = ["Financial Metric", "Value", "Unit", "Notes"]
    for col_idx, h in enumerate(headers_t2, start=2): # Col B to E
        cell = ws1.cell(row=22, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t2_data = [
        ("Documented Placements (Placement Table)", 38, "Placements", "37 have onboard dates"),
        ("Filled Job Openings (Jobs Status)", 57, "Jobs", "51 Filled + 6 Filled Out of Warranty"),
        ("Total Placed Deal Value (VND)", 4351478357, "VND", "Selling price tracked in Placement table"),
        ("Total Placed Deal Value (USD)", 167224.22, "USD", "Equivalent value in USD (38 placements)"),
        ("Placements in 2024 (Onboard Date)", 4, "Placements", "Based on onboard date mapping"),
        ("Placements in 2025 (Onboard Date)", 25, "Placements", "Based on onboard date mapping"),
        ("Placements in 2026 (Onboard Date)", 8, "Placements", "Based on onboard date mapping"),
        ("Placements (Unknown Onboard Date)", 1, "Placements", "Onboard Date missing")
    ]
    
    for row_offset, (metric, val, unit, note) in enumerate(t2_data):
        r = 23 + row_offset
        ws1.row_dimensions[r].height = 18
        c1 = ws1.cell(row=r, column=2, value=metric)
        c2 = ws1.cell(row=r, column=3, value=val)
        c3 = ws1.cell(row=r, column=4, value=unit)
        c4 = ws1.cell(row=r, column=5, value=note)
        
        for c in [c1, c3, c4]:
            c.font = font_regular
            c.border = border_all
        c1.alignment = align_left
        c3.alignment = align_center
        c4.alignment = align_left
        
        c2.font = font_regular
        c2.border = border_all
        c2.alignment = align_right
        
        if "VND" in metric:
            c2.number_format = '#,##0'
        elif "USD" in metric:
            c2.number_format = '$#,##0.00'
        else:
            c2.number_format = '#,##0'
            
        if row_offset % 2 == 1:
            for c in [c1, c2, c3, c4]:
                c.fill = fill_stripe
                
    # Write Table 1.3: Client Status Breakdown (G5:I12)
    ws1["G5"] = "Table 1.3: Client Engagement Status Breakdown"
    ws1["G5"].font = font_section
    
    headers_t3 = ["Account Status", "Client Count", "Commercial Relevance"]
    for col_idx, h in enumerate(headers_t3, start=7): # Col G to I
        cell = ws1.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t3_data = [
        ("Current Client", 18, "Active clients generating jobs"),
        ("Active Opportunity", 10, "Open leads under negotiation"),
        ("Inactive Client", 23, "Temporarily dormant accounts"),
        ("Dead Opportunity", 12, "Unsuccessful client pitches"),
        ("Drop and Stop", 5, "Terminated client relationships"),
        ("Dead Client", 4, "Permanently lost accounts"),
    ]
    
    for row_offset, (status, count, relevance) in enumerate(t3_data):
        r = 7 + row_offset
        ws1.row_dimensions[r].height = 18
        c1 = ws1.cell(row=r, column=7, value=status)
        c2 = ws1.cell(row=r, column=8, value=count)
        c3 = ws1.cell(row=r, column=9, value=relevance)
        
        for c in [c1, c3]:
            c.font = font_regular
            c.border = border_all
        c1.alignment = align_left
        c3.alignment = align_left
        
        c2.font = font_regular
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '#,##0'
        
        if row_offset % 2 == 1:
            for c in [c1, c2, c3]:
                c.fill = fill_stripe
                
    # Total Clients Row
    r = 13
    ws1.row_dimensions[r].height = 18
    c1 = ws1.cell(row=r, column=7, value="Total Clients")
    c1.font = font_bold
    c1.alignment = align_left
    c1.border = border_bottom_double
    c1.fill = fill_accent
    
    c2 = ws1.cell(row=r, column=8, value=72)
    c2.font = font_bold
    c2.alignment = align_right
    c2.border = border_bottom_double
    c2.number_format = '#,##0'
    c2.fill = fill_accent
    
    c3 = ws1.cell(row=r, column=9, value="Total tracked database records")
    c3.font = font_bold
    c3.alignment = align_left
    c3.border = border_bottom_double
    c3.fill = fill_accent

    # Write Table 1.4: Candidate Database Completeness (G16:I26)
    ws1["G16"] = "Table 1.4: Candidate Database Completeness (Base: 810 Profiles)"
    ws1["G16"].font = font_section
    
    headers_t4 = ["Profile Field", "Completed Records", "Completeness Rate"]
    for col_idx, h in enumerate(headers_t4, start=7): # Col G to I
        cell = ws1.cell(row=17, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t4_data = [
        ("CV Link", 808, 0.998),
        ("Summary Note", 746, 0.921),
        ("Email", 645, 0.796),
        ("Phone", 585, 0.722),
        ("LinkedIn URL", 427, 0.527),
        ("CV + At least one contact", 646, 0.798),
        ("Email + Phone + LinkedIn (Full)", 411, 0.507),
        ("Salary Expectation (Min/Max/Currency)", 501, 0.619),
        ("Linked Screening Record", 209, 0.258),
        ("AI Candidate Evaluation Notes", 122, 0.151)
    ]
    
    for row_offset, (field, count, rate) in enumerate(t4_data):
        r = 18 + row_offset
        ws1.row_dimensions[r].height = 18
        c1 = ws1.cell(row=r, column=7, value=field)
        c2 = ws1.cell(row=r, column=8, value=count)
        c3 = ws1.cell(row=r, column=9, value=rate)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        c2.font = font_regular
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '#,##0'
        
        c3.font = font_regular
        c3.alignment = align_right
        c3.border = border_all
        c3.number_format = '0.0%'
        
        if row_offset % 2 == 1:
            c1.fill = fill_stripe
            c2.fill = fill_stripe
            c3.fill = fill_stripe

    # Write Table 1.5: Database Limitations & Missing Fields (B35:E41)
    ws1["B34"] = "Table 1.5: OAC Database Limitations & Missing Fields (Diligence Warnings)"
    ws1["B34"].font = font_section
    
    headers_t5 = ["Missing / Unverifiable Metric", "Airtable Data Status", "Consequence / Limitation", "Alternative Pitch-Safe Proxy"]
    for col_idx, h in enumerate(headers_t5, start=2): # Col B to E
        cell = ws1.cell(row=35, column=col_idx, value=h)
        cell.font = font_header_warn
        cell.fill = fill_header_warn
        cell.alignment = align_header
        cell.border = border_all
        
    t5_data = [
        ("Gross & Operating Profit Trend", "No Financial Records", "KPI table actuals are 0. COGS, commissions, and opex not tracked.", "Use Selling Value Deal-Size tracked in Placement as revenue proxy."),
        ("Academy / Recruiter Trained count", "No Cohort/External database", "Screening calls contain internal training logs only.", "Do not pitch community headcount. Use 107 internal training logs."),
        ("Alumni Network active pool", "No Alumni Table", "No status flags or last touchpoint dates for alumni.", "Use 4,633 raw candidate/application records across Zoho/Airtable."),
        ("90-day Warranty & Replacement rate", "No active warranty flags", "Warranty dates exist but pass/fail rates are unpopulated.", "Use pre-hire failure funnel (245 failed applications) as funnel loss proof."),
        ("Contracting / Project Staffing details", "Under-documented", "Only 3 clients have Contract Dates. Headcount & durations are blank.", "Use placement offer letters, order URLs, invoices in Placement tables."),
        ("Retained Search Track Record", "No retained search tag", "Job Category lists only New Hire, Replacement, Confidential.", "Pitch regional placement/job coverage across 12 countries.")
    ]
    
    for row_offset, (metric, status, consequence, alt) in enumerate(t5_data):
        r = 36 + row_offset
        ws1.row_dimensions[r].height = 28 # higher to wrap text
        c1 = ws1.cell(row=r, column=2, value=metric)
        c2 = ws1.cell(row=r, column=3, value=status)
        c3 = ws1.cell(row=r, column=4, value=consequence)
        c4 = ws1.cell(row=r, column=5, value=alt)
        
        for c in [c1, c2, c3, c4]:
            c.font = font_regular
            c.border = border_all
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
        c2.font = font_bold
        
        if row_offset % 2 == 1:
            for c in [c1, c2, c3, c4]:
                c.fill = fill_stripe

    # Autofit Column Widths Sheet 1
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1: # ignore title row length
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
    # Set manual wrap widths for Table 1.5
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 32

    # ----------------------------------------------------
    # SHEET 2: Pitch_Deck_Metrics_Ratio
    # ----------------------------------------------------
    ws2.row_dimensions[1].height = 25
    ws2.row_dimensions[2].height = 18
    ws2.row_dimensions[3].height = 10
    
    # Title Block
    ws2["B1"] = "OAC Pitch Deck Recruiting Ratios & Metrics"
    ws2["B1"].font = font_title
    ws2["B2"] = "Calculated & Normalized Recruiting Performance Indicators | Language: English"
    ws2["B2"].font = font_italic

    # Table 2.1: Operational Speed & Time-to-Fill (B5:E9)
    ws2["B5"] = "Table 2.1: Operational Speed & Sourcing Funnel Timelines"
    ws2["B5"].font = font_section
    
    headers_t21 = ["Sourcing Funnel Milestone", "Sample Size", "Median (Days)", "Average (Days)"]
    for col_idx, h in enumerate(headers_t21, start=2): # Col B to E
        cell = ws2.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t21_data = [
        ("Time to 1st Source (Job Created -> 1st Candidate)", 71, 4.0, 16.0),
        ("Time to 1st CV Submitted (Job Created -> 1st Sent-out)", 64, 8.5, 22.5),
        ("Time to 1st Interview (Job Created -> 1st Client Interview)", 42, 18.0, 28.0),
        ("Time to Placement (JD Created -> Onboard Date)", 30, 93.5, 128.6)
    ]
    
    for row_offset, (milestone, sample, median, avg) in enumerate(t21_data):
        r = 7 + row_offset
        ws2.row_dimensions[r].height = 18
        c1 = ws2.cell(row=r, column=2, value=milestone)
        c2 = ws2.cell(row=r, column=3, value=sample)
        c3 = ws2.cell(row=r, column=4, value=median)
        c4 = ws2.cell(row=r, column=5, value=avg)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        for c in [c2, c3, c4]:
            c.font = font_regular
            c.alignment = align_right
            c.border = border_all
            
        c2.number_format = '#,##0'
        c3.number_format = '0.0'
        c4.number_format = '0.0'
        
        if row_offset % 2 == 1:
            for c in [c1, c2, c3, c4]:
                c.fill = fill_stripe

    # Table 2.2: Sourcing Funnel Conversion Ratios (B12:E17)
    ws2["B12"] = "Table 2.2: Sourcing Funnel Conversion Rates (Base: 464 Applications)"
    ws2["B12"].font = font_section
    
    headers_t22 = ["Funnel Stage", "Candidate Count", "Conversion Rate", "Strategic Wording for Pitch Deck"]
    for col_idx, h in enumerate(headers_t22, start=2): # Col B to E
        cell = ws2.cell(row=13, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t22_data = [
        ("CV Submitted / Sent-out to Client", 464, 1.000, "464 Verified candidate submissions to clients"),
        ("1st Interview Achieved", 155, 0.334, "33.4% Submission-to-Interview Conversion Rate"),
        ("Final Placements (Conservative)", 38, 0.082, "8.2% Placement Rate based on Placement ledger"),
        ("Final Placements (Broader ATS status)", 57, 0.123, "12.3% Placed Job Opening Rate tracked in ATS"),
        ("Active Late-Stage Pipeline", 21, 0.045, "21 Candidates in Offer/Onboard pipeline (Current Month)")
    ]
    
    for row_offset, (stage, count, conversion, wording) in enumerate(t22_data):
        r = 14 + row_offset
        ws2.row_dimensions[r].height = 18
        c1 = ws2.cell(row=r, column=2, value=stage)
        c2 = ws2.cell(row=r, column=3, value=count)
        c3 = ws2.cell(row=r, column=4, value=conversion)
        c4 = ws2.cell(row=r, column=5, value=wording)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        c2.font = font_regular
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '#,##0'
        
        c3.font = font_regular
        c3.alignment = align_right
        c3.border = border_all
        c3.number_format = '0.0%'
        
        c4.font = font_regular
        c4.alignment = align_left
        c4.border = border_all
        
        if row_offset % 2 == 1:
            for c in [c1, c2, c3, c4]:
                c.fill = fill_stripe

    # Table 2.3: Client Retention & Focus Metrics (G5:I10)
    ws2["G5"] = "Table 2.3: Client Account Stickiness & Concentration"
    ws2["G5"].font = font_section
    
    headers_t23 = ["Retention / Concentration Metric", "Value / Rate", "Formula / Context"]
    for col_idx, h in enumerate(headers_t23, start=7): # Col G to I
        cell = ws2.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t23_data = [
        ("Client Repeat Mandate Rate (2+ jobs)", 0.697, "46 repeat clients / 66 clients with linked jobs"),
        ("Client High-Volume Rate (5+ jobs)", 0.333, "22 high-volume clients / 66 clients with linked jobs"),
        ("Top 1 Client Share (RightShip)", 0.252, "RightShip represents 25.2% of placement deal value"),
        ("Top 3 Client Concentration", 0.608, "RightShip, Omnistream, StoryCO represent 60.8% of value"),
        ("Top 5 Client Concentration", 0.784, "Top 5 clients represent 78.4% of total deal value")
    ]
    
    for row_offset, (metric, val, formula) in enumerate(t23_data):
        r = 7 + row_offset
        ws2.row_dimensions[r].height = 18
        c1 = ws2.cell(row=r, column=7, value=metric)
        c2 = ws2.cell(row=r, column=8, value=val)
        c3 = ws2.cell(row=r, column=9, value=formula)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        c2.font = font_bold
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '0.0%'
        
        c3.font = font_regular
        c3.alignment = align_left
        c3.border = border_all
        
        if row_offset % 2 == 1:
            for c in [c1, c2, c3]:
                c.fill = fill_stripe

    # Table 2.4: Geographic Placement Value (G13:I18)
    ws2["G12"] = "Table 2.4: Geographic Breakdown of Placed Revenue Value"
    ws2["G12"].font = font_section
    
    headers_t24 = ["Placement Country", "Tracked Placement (VND)", "Market Share"]
    for col_idx, h in enumerate(headers_t24, start=7): # Col G to I
        cell = ws2.cell(row=13, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t24_data = [
        ("Vietnam", 2951000000, 0.678),
        ("Australia", 1008000000, 0.232),
        ("Singapore", 214900000, 0.049),
        ("Japan", 90200000, 0.021),
        ("Canada", 86700000, 0.020)
    ]
    
    for row_offset, (country, vnd, share) in enumerate(t24_data):
        r = 14 + row_offset
        ws2.row_dimensions[r].height = 18
        c1 = ws2.cell(row=r, column=7, value=country)
        c2 = ws2.cell(row=r, column=8, value=vnd)
        c3 = ws2.cell(row=r, column=9, value=share)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        c2.font = font_regular
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '#,##0'
        
        c3.font = font_regular
        c3.alignment = align_right
        c3.border = border_all
        c3.number_format = '0.0%'
        
        if row_offset % 2 == 1:
            for c in [c1, c2, c3]:
                c.fill = fill_stripe
                
    # Total Geo Row
    r = 19
    ws2.row_dimensions[r].height = 18
    c1 = ws2.cell(row=r, column=7, value="Total Geographies")
    c1.font = font_bold
    c1.alignment = align_left
    c1.border = border_bottom_double
    c1.fill = fill_accent
    
    c2 = ws2.cell(row=r, column=8, value=4350800000)
    c2.font = font_bold
    c2.alignment = align_right
    c2.border = border_bottom_double
    c2.number_format = '#,##0'
    c2.fill = fill_accent
    
    c3 = ws2.cell(row=r, column=9, value=1.000)
    c3.font = font_bold
    c3.alignment = align_right
    c3.border = border_bottom_double
    c3.number_format = '0.0%'
    c3.fill = fill_accent

    # Table 2.5: Sourcing Candidate Acquisition Channels (B22:D27)
    ws2["B21"] = "Table 2.5: Candidate Acquisition Channels (Base: 810 Candidates)"
    ws2["B21"].font = font_section
    
    headers_t25 = ["Acquisition Channel", "Profiles Acquired", "Percentage Share"]
    for col_idx, h in enumerate(headers_t25, start=2): # Col B to D
        cell = ws2.cell(row=22, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t25_data = [
        ("LinkedIn Sourcing / InMail", 420, 0.519),
        ("Facebook Sourcing / Community", 232, 0.286),
        ("Other / Organic Job boards", 129, 0.159),
        ("Referral Network", 25, 0.031),
        ("Google Organic Search", 2, 0.002)
    ]
    
    for row_offset, (channel, count, share) in enumerate(t25_data):
        r = 23 + row_offset
        ws2.row_dimensions[r].height = 18
        c1 = ws2.cell(row=r, column=2, value=channel)
        c2 = ws2.cell(row=r, column=3, value=count)
        c3 = ws2.cell(row=r, column=4, value=share)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        c2.font = font_regular
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '#,##0'
        
        c3.font = font_regular
        c3.alignment = align_right
        c3.border = border_all
        c3.number_format = '0.0%'
        
        if row_offset % 2 == 1:
            c1.fill = fill_stripe
            c2.fill = fill_stripe
            c3.fill = fill_stripe
            
    # Total Sourcing Channels Row
    r = 28
    ws2.row_dimensions[r].height = 18
    c1 = ws2.cell(row=r, column=2, value="Total Structured Pool")
    c1.font = font_bold
    c1.alignment = align_left
    c1.border = border_bottom_double
    c1.fill = fill_accent
    
    c2 = ws2.cell(row=r, column=3, value=810)
    c2.font = font_bold
    c2.alignment = align_right
    c2.border = border_bottom_double
    c2.number_format = '#,##0'
    c2.fill = fill_accent
    
    c3 = ws2.cell(row=r, column=4, value=1.000)
    c3.font = font_bold
    c3.alignment = align_right
    c3.border = border_bottom_double
    c3.number_format = '0.0%'
    c3.fill = fill_accent

    # Table 2.6: Top 10 Recruiter Selling Value Attribution (G22:I32)
    ws2["G21"] = "Table 2.6: Top 10 Recruiter Performance Selling Value"
    ws2["G21"].font = font_section
    
    headers_t26 = ["Recruiter (System Owner)", "Selling Value (VND)", "Performance Share (%)"]
    for col_idx, h in enumerate(headers_t26, start=7): # Col G to I
        cell = ws2.cell(row=22, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_all
        
    t26_data = [
        ("Peter", 1453000000, 0.334),
        ("Lauren", 883200000, 0.203),
        ("Serena", 763800000, 0.176),
        ("Ellie", 714400000, 0.164),
        ("Vicky", 492500000, 0.113),
        ("Đào Hà", 482100000, 0.111),
        ("Amy", 242700000, 0.056),
        ("Tina", 216000000, 0.050),
        ("Anna", 141600000, 0.033),
        ("Alice", 110200000, 0.025)
    ]
    
    for row_offset, (owner, val, share) in enumerate(t26_data):
        r = 23 + row_offset
        ws2.row_dimensions[r].height = 18
        c1 = ws2.cell(row=r, column=7, value=owner)
        c2 = ws2.cell(row=r, column=8, value=val)
        c3 = ws2.cell(row=r, column=9, value=share)
        
        c1.font = font_regular
        c1.alignment = align_left
        c1.border = border_all
        
        c2.font = font_regular
        c2.alignment = align_right
        c2.border = border_all
        c2.number_format = '#,##0'
        
        c3.font = font_regular
        c3.alignment = align_right
        c3.border = border_all
        c3.number_format = '0.0%'
        
        if row_offset % 2 == 1:
            for c in [c1, c2, c3]:
                c.fill = fill_stripe

    # Autofit Column Widths Sheet 2
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1: # ignore title row length
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws2.column_dimensions["E"].width = 38 # Wording column wrap
    
    # Save Workbook
    wb.save(excel_path)
    print(f"Excel successfully created at: {excel_path}")

if __name__ == "__main__":
    create_oac_excel()
